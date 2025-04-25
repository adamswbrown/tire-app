from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import pandas as pd
import re

# Normalize fields for comparison
def normalize(field):
    return field.strip().lower().replace(" ", "_")

# Load workbook and worksheet
wb = load_workbook("/Users/adambrown/Developer/Altra/TIREApp/Copy of UKPowerNetworks-App to Server Dr Migrate template.xlsx")
ws = wb.active

# Expected field names from SQL + manual field
expected_fields = set([
    "Application Name",  # manually added
    "app_guid", "app_owner", "app_owner_email", "app_function", "app_criticality",
    "business_unit", "app_type", "data_classification", "disaster_recovery",
    "high_availiability", "deployment", "environments", "backup_restore_policies",
    "app_owner_azure_knowledge", "app_sme", "app_sme_email", "app_sme_azure_knowledge",
    "supported_by", "support_team_size", "support_team_azure_knowledge",
    "resilience_data_retention", "health_stability", "health_documentation",
    "app_description", "app_type_cots_software_name", "app_type_cots_vendor_name",
    "app_type_cots_version_installed", "app_type_cots_url", "app_type_cots_can_run_in_azure",
    "app_type_cots_can_modernised_paas", "app_type_cots_notes", "treatment_pattern",
    "payment_model", "pii_data", "business_critical", "backups"
])

# Extract headers from row 4
headers = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=4, column=col).value
    if header:
        headers[col] = header.strip()

# Map validations to column indexes
validation_map = {}
for dv in ws.data_validations.dataValidation:
    for sq in str(dv.sqref).split():
        match = re.match(r"([A-Z]+)", sq)
        if match:
            col_letter = match.group(1)
            col_index = column_index_from_string(col_letter)
            validation_map[col_index] = dv

# Build validation data
validation_data = []

for col_index, header in headers.items():
    dv = validation_map.get(col_index)
    validation_data.append({
        "Header": header,
        "Type": dv.type if dv else "No Validation",
        "Values": dv.formula1 if dv else "—",
        "Allow Blank": dv.allow_blank if dv else "—",
        "Show Error": dv.showErrorMessage if dv else "—"
    })

# Create DataFrame
df_validations = pd.DataFrame(validation_data)

# Normalize field names for comparison
normalized_sheet_fields = set(normalize(h) for h in df_validations["Header"])
normalized_expected_fields = set(normalize(f) for f in expected_fields)

# Compare
missing_fields = normalized_expected_fields - normalized_sheet_fields
extra_fields = normalized_sheet_fields - normalized_expected_fields

# Output table
print("\n=== Validation Table ===")
print(df_validations.to_string(index=False))

# Validation summary
print("\n=== Validation Check ===")
print(f"Total expected fields: {len(expected_fields)}")
print(f"Total found in sheet: {len(headers)}")

if missing_fields:
    print(f"\n⚠️ Missing in sheet ({len(missing_fields)}):")
    for field in sorted(missing_fields):
        print(f" - {field}")
else:
    print("✅ All expected fields are present.")

if extra_fields:
    print(f"\n⚠️ Extra fields in sheet ({len(extra_fields)}):")
    for field in sorted(extra_fields):
        print(f" - {field}")
else:
    print("✅ No unexpected fields found.")